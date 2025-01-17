# This script was developed by SMC LiverAIz Soyoung Lim with CursorAI
# 2025.01 (vers.1)

import os
import nibabel as nib
import pandas as pd
import numpy as np
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm
from openpyxl.styles import PatternFill
import openpyxl
import time
from datetime import datetime
from multiprocessing import Pool, cpu_count
import mmap
import functools
import concurrent.futures

def is_mask_empty(mask_data):
    # 메모리 효율을 위해 청크 단위로 처리
    chunk_size = 10
    for i in range(0, mask_data.shape[2], chunk_size):
        chunk = mask_data.dataobj[..., i:min(i+chunk_size, mask_data.shape[2])]
        if np.any(chunk):
            return False
    return True

def process_single_case(args):
    img_file, img_path, mask_folder, mask_names = args
    results = []
    
    case_id = img_file.split('.')[0]
    img_data = nib.load(os.path.join(img_path, img_file))
    img_shape = img_data.shape
    img_affine = img_data.affine
    
    img_info = {
        'case_id': case_id,
        'dimensions': str(img_shape),
        'origin': str(img_affine[0:3, 3]),
        'spacing': str(np.sqrt(np.sum(img_affine[:3, :3]**2, axis=0)))
    }
    
    mask_info = {
        'case_id': case_id,
        **{name: 0 for name in mask_names}
    }
    case_mask_folder = os.path.join(mask_folder, case_id)
    
    if os.path.exists(case_mask_folder):
        for mask_file in os.listdir(case_mask_folder):
            if mask_file.endswith('.nii.gz'):
                mask_name = mask_file.split('.')[0]
                mask_data = nib.load(os.path.join(case_mask_folder, mask_file))
                
                if not is_mask_empty(mask_data):
                    mask_info[mask_name] = 1
                
                results.append({
                    'case_id': case_id,
                    'mask_name': mask_name,
                    'dimensions_match': img_shape == mask_data.shape,
                    'origin_match': np.allclose(img_affine, mask_data.affine)
                })
    
    return results, mask_info, img_info

def update_excel_with_mask_presence(reference_data, df_mask_presence, mask_names, case_id_col):
    # 엑셀 파일 로드
    workbook = openpyxl.load_workbook('data.xlsx')
    sheet = workbook.active
    
    # 열 인덱스 찾기
    header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col_indices = {}
    
    # case_id 열 찾기
    for idx, cell in enumerate(header_row, 1):
        if cell and isinstance(cell, str):
            if 'case' in cell.lower() and 'id' in cell.lower():
                col_indices['case_id'] = idx
            # 마스크 열 찾기
            for mask_name in mask_names:
                if mask_name.lower() in cell.lower():
                    col_indices[mask_name] = idx
    
    # 노란색 배경 스타일 정의
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    # 각 행 업데이트
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), 2):
        case_id = str(row[col_indices['case_id']-1].value)
        mask_row = df_mask_presence[df_mask_presence['case_id'] == case_id]
        
        if not mask_row.empty:
            for mask_name in mask_names:
                if mask_name in col_indices:
                    col_idx = col_indices[mask_name]
                    current_cell = sheet.cell(row=row_idx, column=col_idx)
                    new_value = mask_row.iloc[0][mask_name]
                    
                    # 현재 값을 숫자로 변환
                    current_value = None
                    if current_cell.value is not None:
                        if isinstance(current_cell.value, (int, float)):
                            current_value = int(current_cell.value)
                        elif isinstance(current_cell.value, str):
                            if current_cell.value.lower() in ['o', 'y', 'yes', '1']:
                                current_value = 1
                            elif current_cell.value.lower() in ['x', 'n', 'no', '0']:
                                current_value = 0
                    
                    # 값이 다른 경우 업데이트하고 노란색으로 표시
                    if current_value != new_value:
                        current_cell.value = new_value
                        current_cell.fill = yellow_fill
    
    # 변경사항 저장
    workbook.save('data_updated.xlsx')
    print("엑셀 파일이 업데이트되었습니다. (data_updated.xlsx)")

def check_volume_info(img_path, mask_folder):
    start_time = time.time()
    print(f"처리 시작 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # 마스크 이름 리스트 생성
    first_case = next(os.walk(mask_folder))[1][0]
    mask_names = [f.split('.')[0] for f in os.listdir(os.path.join(mask_folder, first_case)) 
                 if f.endswith('.nii.gz')]
    mask_names.sort()
    
    # 처리할 파일 목록 생성
    img_files = [f for f in os.listdir(img_path) if f.endswith('.nii.gz')]
    args_list = [(f, img_path, mask_folder, mask_names) for f in img_files]
    
    all_results = []
    all_mask_info = []
    all_img_info = []
    
    # 멀티스레딩 처리
    with ThreadPoolExecutor(max_workers=os.cpu_count()) as executor:
        futures = [executor.submit(process_single_case, args) for args in args_list]
        
        # tqdm으로 진행률 표시
        for future in tqdm(as_completed(futures), total=len(futures), desc="처리 중"):
            results, mask_info, img_info = future.result()
            all_results.extend(results)
            all_mask_info.append(mask_info)
            all_img_info.append(img_info)
    
    # DataFrame 생성 및 비교 처리
    df_results = pd.DataFrame(all_results)
    df_mask_presence = pd.DataFrame(all_mask_info)
    df_img_info = pd.DataFrame(all_img_info)
    
    # data.xlsx 처리
    reference_data = pd.read_excel('data.xlsx')
    
    # case_id 열 이름 찾기 (대소문자 무시)
    case_id_col = None
    for col in reference_data.columns:
        if 'case' in col.lower() and 'id' in col.lower():
            case_id_col = col
            break
    
    if case_id_col is None:
        print("경고: data.xlsx에서 case id 열을 찾을 수 없습니다.")
        case_id_col = 'case_id'  # 기본값 설정
    
    # case_id를 문자열로 변환하여 통일
    reference_data[case_id_col] = reference_data[case_id_col].astype(str)
    df_mask_presence['case_id'] = df_mask_presence['case_id'].astype(str)
    
    comparison_results = []
    
    # 각 케이스와 마스크에 대해 비교
    for case_id in df_mask_presence['case_id']:
        mask_row = df_mask_presence[df_mask_presence['case_id'] == case_id].iloc[0]
        ref_case = reference_data[reference_data[case_id_col] == case_id]
        
        if not ref_case.empty:
            ref_row = ref_case.iloc[0]
            
            for mask_name in mask_names:
                # data.xlsx에서 해당 마스크 열 찾기 (대소문자 무시)
                mask_col = None
                for col in reference_data.columns:
                    if mask_name.lower() in col.lower():
                        mask_col = col
                        break
                
                excel_value = 'Column Not Found'
                if mask_col:
                    excel_value = ref_row[mask_col]
                    # 숫자가 아닌 경우 처리 (예: 'O', 'X', 'Y', 'N' 등)
                    if isinstance(excel_value, str):
                        if excel_value.lower() in ['o', 'y', 'yes', '1']:
                            excel_value = 1
                        elif excel_value.lower() in ['x', 'n', 'no', '0']:
                            excel_value = 0
                
                comparison_results.append({
                    'case_id': case_id,
                    'mask_name': mask_name,
                    'presence_in_folder': mask_row[mask_name],
                    'presence_in_excel': excel_value,
                    'match': mask_row[mask_name] == excel_value if excel_value not in ['Column Not Found'] else False
                })
        else:
            # data.xlsx에서 케이스를 찾을 수 없는 경우
            for mask_name in mask_names:
                comparison_results.append({
                    'case_id': case_id,
                    'mask_name': mask_name,
                    'presence_in_folder': mask_row[mask_name],
                    'presence_in_excel': 'Case Not Found',
                    'match': False
                })
    
    df_comparison = pd.DataFrame(comparison_results)
    
    # data.xlsx 업데이트
    update_excel_with_mask_presence(reference_data, df_mask_presence, mask_names, case_id_col)
    
    # 결과 저장
    with pd.ExcelWriter('volume_analysis_results.xlsx', engine='openpyxl') as writer:
        df_results.to_excel(writer, sheet_name='Volume Comparison', index=False)
        df_mask_presence.to_excel(writer, sheet_name='Mask Presence', index=False)
        df_img_info.to_excel(writer, sheet_name='Image Info', index=False)
        df_comparison.to_excel(writer, sheet_name='Mask Presence Comparison', index=False)
    
    end_time = time.time()
    elapsed_time = end_time - start_time
    minutes = int(elapsed_time // 60)
    seconds = int(elapsed_time % 60)
    
    print(f"\n처리 완료 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"총 소요 시간: {minutes}분 {seconds}초")
    print(f"처리된 케이스 수: {len(img_files)}")
    print(f"케이스당 평균 처리 시간: {elapsed_time/len(img_files):.2f}초")

# 메인 실행 부분을 if __name__ == '__main__': 블록으로 감싸기
if __name__ == '__main__':
    # 실행 경로 설정
    img_path = 'img'
    mask_folder = 'mask'
    
    # 볼륨 정보 체크 실행
    check_volume_info(img_path, mask_folder)